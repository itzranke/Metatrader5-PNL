"""Test Phase 8 — Performance score, deposit/withdrawal, excel export."""
import io
from datetime import UTC

from openpyxl import load_workbook

from packages.analytics import performance_score


def _setup(client, username="p8user", email="p8@example.com"):
    client.post(
        "/api/v1/auth/register",
        json={"username": username, "email": email, "password": "rahasia123"},
    )
    token = client.post(
        "/api/v1/auth/login", json={"email": email, "password": "rahasia123"}
    ).json()["access_token"]
    h = {"Authorization": f"Bearer {token}"}
    acc = client.post(
        "/api/v1/accounts", json={"kind": "demo", "name": "Data Contoh"}, headers=h
    ).json()
    return token, h, acc


# ---------------------------------------------------------------- unit analytics


def _t(net, r=None, mae=None, mfe=None, day=1):
    from datetime import datetime, timedelta

    now = datetime.now(UTC)
    return {
        "net_profit": net, "gross_profit": net if net > 0 else 0,
        "open_time": now - timedelta(days=day, hours=2),
        "close_time": now - timedelta(days=day),
        "r_multiple": r, "mae": mae, "mfe": mfe,
    }


def test_score_requires_20_trades():
    result = performance_score([_t(10)] * 10)
    assert result["score"] is None
    assert result["progress"] == 10 and result["need"] == 20


def test_score_winning_trader_high():
    trades = [_t(100, r=1.5, mae=-20, mfe=150, day=i) for i in range(1, 25)]
    result = performance_score(
        trades,
        plan_match_rate=0.8, rule_adherence_rate=0.9,
        revenge_ratio=0.0, emotion_stability=0.8, journal_count=12,
    )
    assert result["score"] is not None
    assert 0 <= result["score"] <= 100
    assert result["score"] >= 70  # trader konsisten profit → Strong/Excellent
    assert result["label"] in ("Strong", "Excellent")
    assert result["data_complete"] is True


def test_score_losing_trader_low():
    trades = [_t(-50, r=-1.8, mae=-80, mfe=30, day=i) for i in range(1, 25)]
    result = performance_score(trades, journal_count=0)
    assert result["score"] is not None
    assert result["score"] < 55  # rugi konsisten → tidak Good
    assert result["data_complete"] is False  # jurnal kurang → penalti data


def test_score_all_components_present():
    trades = [_t(50, r=1.2, mae=-10, mfe=120, day=i) for i in range(1, 25)]
    result = performance_score(trades, journal_count=12)
    assert set(result["components"]) == {"risk_mgmt", "consistency", "profitability", "drawdown", "trade_quality", "discipline"}


# ---------------------------------------------------------------- API


def test_score_endpoint_demo_account(client):
    token, h, acc = _setup(client)
    r = client.get(f"/api/v1/accounts/{acc['id']}/score", headers=h)
    assert r.status_code == 200
    body = r.json()
    assert body["score"] is not None and 0 <= body["score"] <= 100
    assert body["progress"] >= 120
    assert body["label"] is not None


def test_score_cross_user_404(client):
    token_a, h_a, acc_a = _setup(client)
    client.post(
        "/api/v1/auth/register",
        json={"username": "p8b", "email": "p8b@example.com", "password": "rahasia123"},
    )
    token_b = client.post(
        "/api/v1/auth/login", json={"email": "p8b@example.com", "password": "rahasia123"}
    ).json()["access_token"]
    r = client.get(
        f"/api/v1/accounts/{acc_a['id']}/score",
        headers={"Authorization": f"Bearer {token_b}"},
    )
    assert r.status_code == 404


def test_deposit_withdrawal_crud(client):
    token, h, acc = _setup(client)
    m0_body = client.get(f"/api/v1/accounts/{acc['id']}/money", headers=h).json()
    m0 = m0_body["net_deposits"]
    m0_items = m0_body["items"]

    r = client.post(
        f"/api/v1/accounts/{acc['id']}/deposits",
        json={"amount": 500.0, "method": "bank", "note": "Top up"},
        headers=h,
    )
    assert r.status_code == 201
    dep = r.json()
    assert dep["kind"] == "deposit" and dep["amount"] == 500.0

    r2 = client.post(
        f"/api/v1/accounts/{acc['id']}/withdrawals",
        json={"amount": 120.5, "method": "bank", "note": "Tarik"},
        headers=h,
    )
    assert r2.status_code == 201

    r3 = client.get(f"/api/v1/accounts/{acc['id']}/money", headers=h)
    assert r3.status_code == 200
    body = r3.json()
    # demo generator acak (withdrawal tidak selalu ada) → bandingkan delta saja
    assert body["net_deposits"] == m0 + 500.0 - 120.5
    assert len(body["items"]) == len(m0_items) + 2
    kinds = {i["kind"] for i in body["items"]}
    assert kinds == {"deposit", "withdrawal"}

    # hapus
    assert client.delete(f"/api/v1/money/deposit/{dep['id']}", headers=h).status_code == 204
    r4 = client.get(f"/api/v1/money/deposit/{dep['id']}", headers=h)
    assert r4.status_code in (404, 405)  # endpoint GET tidak ada → 405/404


def test_money_cross_user_404(client):
    token_a, h_a, acc_a = _setup(client)
    client.post(
        "/api/v1/auth/register",
        json={"username": "p8c", "email": "p8c@example.com", "password": "rahasia123"},
    )
    token_b = client.post(
        "/api/v1/auth/login", json={"email": "p8c@example.com", "password": "rahasia123"}
    ).json()["access_token"]
    r = client.post(
        f"/api/v1/accounts/{acc_a['id']}/deposits",
        json={"amount": 10},
        headers={"Authorization": f"Bearer {token_b}"},
    )
    assert r.status_code == 404


def test_excel_export_multisheet(client):
    token, h, acc = _setup(client)
    r = client.get(f"/api/v1/accounts/{acc['id']}/export/excel.xlsx", headers=h)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    assert set(wb.sheetnames) == {"Trades", "Journal", "Metrik"}
    assert wb["Trades"].max_row >= 121  # header + ≥120 trade
    assert wb["Metrik"].max_row >= 3  # header + metrik
